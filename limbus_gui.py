# -*- coding: utf-8 -*-
"""
边狱巴士 人格DPS计算器 GUI 版
tkinter 实现：自选人格/技能、环境参数、期望伤害表、拼点模拟、Excel导出、人格数据管理
"""
import json, os, sys, math, random, itertools, subprocess, tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

# ---------- 数据文件定位（兼容 PyInstaller 打包） ----------
def data_path(name):
    if getattr(sys, "frozen", False):
        # 优先读 exe 同目录的可编辑副本，否则用打包内置的
        local = os.path.join(os.path.dirname(sys.executable), name)
        if os.path.exists(local): return local
        return os.path.join(sys._MEIPASS, name)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), name)

IDENT_FILE = data_path("identities.json")

CFG = {
    "sanity_to_heads": 0.01, "heads_min": 0.05, "heads_max": 0.95,
    "crit_chance": 0.05, "crit_mult": 1.2, "level_diff_coef": 0.03,
    "resist_table": {"fatal": 1.5, "normal": 1.0, "endured": 0.5, "ineffective": 0.0},
}
RESIST_CN = {"fatal": "致命(1.5x)", "normal": "普通(1.0x)",
             "endured": "耐受(0.5x)", "ineffective": "无效(0x)"}
TYPE_CN = {"slash": "斩击", "pierce": "突刺", "blunt": "打击"}
SIN_CN = {"wrath":"愤怒","lust":"色欲","sloth":"怠惰","gluttony":"暴食",
          "gloom":"忧郁","pride":"傲慢","envy":"嫉妒"}

# ---------- 计算引擎（与 CLI 版一致） ----------
def heads_prob(sanity):
    return min(max(0.5 + sanity * CFG["sanity_to_heads"], CFG["heads_min"]), CFG["heads_max"])

def resolve_cond(skill, charge):
    base = skill["base_power"]; coins = list(skill["coin_power"]); tag = ""
    c = skill.get("cond")
    if c and charge >= c.get("charge_at_least", 10**9):
        base += c.get("base_power_add", 0)
        coins = [x + c.get("coin_power_add", 0) for x in coins]
        tag = f"充能≥{c['charge_at_least']}已触发"
    return base, coins, tag

def damage_multiplier(phys, sin_res, off_lv, def_lv, power_up, dmg_up, fragile):
    m = CFG["resist_table"][phys] * CFG["resist_table"][sin_res]
    m *= 1 + CFG["level_diff_coef"] * (off_lv - def_lv)
    return m * (1 + power_up/100) * (1 + dmg_up/100) * (1 + fragile/100)

def skill_damage_dist(base, coins, p, mult, cc, cm):
    dist = {}
    for outcomes in itertools.product([0, 1], repeat=len(coins)):
        prob = 1.0; total = 0.0
        for cp, h in zip(coins, outcomes):
            prob *= p if h else (1 - p)
            hit = base + (cp if h else 0)
            hit *= 1 + cc * (cm - 1)
            total += hit
        total = round(total * mult, 2)
        dist[total] = dist.get(total, 0) + prob
    return sorted(dist.items())

def summarize(dist):
    ev = sum(d*q for d, q in dist)
    sd = math.sqrt(sum((d-ev)**2*q for d, q in dist))
    return ev, sd, dist[0][0], dist[-1][0]

def clash_sim(a, b, p_a, p_b, rounds=10000):
    a_base, a_coins = a; b_base, b_coins = b
    win = 0; remain = 0
    for _ in range(rounds):
        ca = list(a_coins); cb = list(b_coins)
        while ca and cb:
            pa = a_base + sum(cp for cp in ca if random.random() < p_a)
            pb = b_base + sum(cp for cp in cb if random.random() < p_b)
            if pa > pb: cb.pop()
            elif pb > pa: ca.pop()
        if ca: win += 1; remain += len(ca)
    return win/rounds*100, remain/rounds

# ---------- E.G.O 与 DoT ----------
def sin_income(deck):
    """循环内打出每张技能牌产生1个对应罪孽资源"""
    inc = {}
    for s in deck:
        if s.get("is_ego"): continue
        inc[s["sin"]] = inc.get(s["sin"], 0) + s.get("copies", 1)
    return inc

def ego_uses(ego, income, turns=6):
    """资源限制下每循环E.G.O可用次数（且不超过回合数）"""
    cost = ego.get("cost", {})
    if not cost: return 1
    uses = min(income.get(s, 0) // n for s, n in cost.items())
    return min(uses, turns)

DOT_CN = {"bleed": "流血", "burn": "燃烧", "sinking": "沉潜"}
DOT_CN_REV = {"流血": "bleed", "燃烧": "burn", "沉潜": "sinking"}
SINS_LIST = list(SIN_CN.keys())
def dot_cycle_damage(deck, turns=6, enemy_flips=12):
    """DoT循环伤害估算：
    流血：中招者拼点或使用硬币时触发 → potency × min(count, 敌方每循环硬币数)
    沉潜：中招者被命中时触发 → potency × min(count, 我方循环总命中数)
    燃烧：回合结束触发 → potency × min(count, 回合数)
    按每张技能牌独立赋予计算（上限假设，未考虑层数叠加与抵抗）。"""
    hits = sum(s.get("copies", 1) * len(s["coin_power"]) for s in deck if not s.get("is_ego"))
    total = 0.0; detail = []
    for s in deck:
        dot = s.get("dot")
        if not dot or s.get("is_ego"): continue
        p_, c_ = dot.get("potency", 0), dot.get("count", 0)
        for _ in range(s.get("copies", 1)):
            if dot["type"] == "burn":
                dmg = p_ * min(c_, turns)
            elif dot["type"] == "bleed":
                dmg = p_ * min(c_, enemy_flips)
            else:  # sinking
                dmg = p_ * min(c_, hits)
            total += dmg
        detail.append(f"{s['name']}({DOT_CN.get(dot['type'],dot['type'])} {p_}强{c_}次×{s.get('copies',1)}张)")
    return total, detail, hits


# ---------- GUI ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("边狱巴士 · 人格DPS计算器")
        self.geometry("1080x640")
        self.load_data()
        self.build_ui()

    def load_data(self):
        with open(IDENT_FILE, encoding="utf-8") as f:
            self.data = json.load(f)
        self.idents = self.data["identities"]

    # ---- 布局 ----
    def _safe(self, fn):
        """包装回调：异常弹窗而不是静默失败（exe无控制台，异常不可见）"""
        def wrapper(*a, **k):
            try:
                return fn(*a, **k)
            except Exception as ex:
                import traceback
                messagebox.showerror("操作失败", f"{ex}\n\n{traceback.format_exc()[-500:]}")
        return wrapper

    def build_ui(self):
        top = ttk.Frame(self); top.pack(fill="x", padx=8, pady=6)

        # 人格选择
        ttk.Label(top, text="人格:").grid(row=0, column=0, sticky="w")
        self.ident_var = tk.StringVar()
        self.ident_cb = ttk.Combobox(top, textvariable=self.ident_var, state="readonly", width=28,
                                     values=[i["name"] for i in self.idents])
        self.ident_cb.grid(row=0, column=1, padx=4)
        self.ident_cb.bind("<<ComboboxSelected>>", lambda e: self.refresh_skills())
        self.ident_cb.current(0)

        # 环境参数
        env = ttk.LabelFrame(self, text="环境参数"); env.pack(fill="x", padx=8, pady=4)
        self.vars = {}
        fields = [("理智度", "sanity", 0, -45, 45), ("充能层数", "charge", 0, 0, 30),
                  ("敌方防御等级", "def_lv", 50, 1, 90), ("威力提升%", "power_up", 0, -100, 300),
                  ("伤害提升%", "dmg_up", 0, -100, 300), ("敌方脆弱%", "fragile", 0, 0, 300),
                  ("敌方硬币/循环", "enemy_flips", 12, 0, 60)]
        for col, (label, key, default, lo, hi) in enumerate(fields):
            ttk.Label(env, text=label).grid(row=0, column=col*2, padx=2, sticky="e")
            v = tk.IntVar(value=default); self.vars[key] = v
            ttk.Spinbox(env, from_=lo, to=hi, textvariable=v, width=6).grid(row=0, column=col*2+1, padx=2)

        ttk.Label(env, text="物理抗性").grid(row=1, column=0, padx=2, sticky="e")
        self.phys_var = tk.StringVar(value="normal")
        ttk.Combobox(env, textvariable=self.phys_var, state="readonly", width=12,
                     values=[f"{k} {RESIST_CN[k]}" for k in CFG["resist_table"]]).grid(row=1, column=1)
        ttk.Label(env, text="罪孽抗性").grid(row=1, column=2, padx=2, sticky="e")
        self.sin_var = tk.StringVar(value="normal")
        ttk.Combobox(env, textvariable=self.sin_var, state="readonly", width=12,
                     values=[f"{k} {RESIST_CN[k]}" for k in CFG["resist_table"]]).grid(row=1, column=3)
        ttk.Label(env, text="暴击率").grid(row=1, column=4, padx=2, sticky="e")
        self.crit_var = tk.DoubleVar(value=CFG["crit_chance"])
        ttk.Spinbox(env, from_=0, to=1, increment=0.05, textvariable=self.crit_var, width=6).grid(row=1, column=5)

        # 按钮
        btns = ttk.Frame(self); btns.pack(fill="x", padx=8, pady=4)
        ttk.Button(btns, text="计算伤害期望", command=self._safe(self.calc)).pack(side="left", padx=4)
        ttk.Button(btns, text="拼点模拟(选中两项)", command=self._safe(self.clash)).pack(side="left", padx=4)
        ttk.Button(btns, text="导出Excel", command=self._safe(self.export)).pack(side="left", padx=4)
        ttk.Button(btns, text="自定义技能组", command=self._safe(self.custom_deck_dialog)).pack(side="left", padx=12)
        ttk.Button(btns, text="从Wiki获取人格", command=self._safe(self.wiki_dialog)).pack(side="left", padx=4)
        ttk.Button(btns, text="新增人格", command=self._safe(self.add_identity_dialog)).pack(side="left", padx=4)
        ttk.Button(btns, text="编辑数据文件", command=self._safe(self.open_json)).pack(side="left", padx=4)
        ttk.Button(btns, text="重新加载", command=self._safe(self.reload)).pack(side="left", padx=4)

        # 结果表
        cols = ("skill","type","sin","base","coins","n","copies","ev","sd","range","cycle","tag")
        heads = ["技能","类型","罪孽","基础威力","硬币威力","硬币数","卡组份数","期望伤害","标准差","范围","循环贡献","条件"]
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=15)
        for c, h in zip(cols, heads):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=80 if c not in ("skill","coins","tag") else 140, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=8, pady=4)
        self.tree.tag_configure("odd", background="#F2F6FC")

        self.status = tk.StringVar(value="就绪")
        ttk.Label(self, textvariable=self.status, anchor="w").pack(fill="x", padx=8, pady=2)
        self.refresh_skills()

    # ---- 逻辑 ----
    def current_ident(self):
        return self.idents[self.ident_cb.current()]

    def env(self):
        phys = self.phys_var.get().split()[0]
        sinr = self.sin_var.get().split()[0]
        return {**{k: v.get() for k, v in self.vars.items()},
                "phys": phys, "sin_res": sinr, "crit": self.crit_var.get()}

    def refresh_skills(self):
        self.tree.delete(*self.tree.get_children())
        ident = self.current_ident()
        for idx, sk in enumerate(ident["skills"]):
            self.tree.insert("", "end", values=(
                sk["name"], TYPE_CN[sk["dmg_type"]], SIN_CN[sk["sin"]],
                sk["base_power"], "+".join(map(str, sk["coin_power"])),
                len(sk["coin_power"]), sk.get("copies", {0:3,1:2,2:1}.get(idx,1)), "-", "-", "-", "-", sk.get("cond", "")))
        self.status.set(f"已选择：{ident['name']}（{ident.get('note','')}）— 点「计算伤害期望」")

    def calc(self):
        e = self.env(); p = heads_prob(e["sanity"])
        ident = self.current_ident()
        self.tree.delete(*self.tree.get_children())
        self._report = []
        cycle_total = 0.0; cycle_cards = 0
        for i, sk in enumerate(ident["skills"]):
            base, coins, tag = resolve_cond(sk, e["charge"])
            mult = damage_multiplier(e["phys"], e["sin_res"], ident.get("offense_level", 50),
                                     e["def_lv"], e["power_up"], e["dmg_up"], e["fragile"])
            ev, sd, lo, hi = summarize(skill_damage_dist(base, coins, p, mult, e["crit"], CFG["crit_mult"]))
            copies = sk.get("copies", {0:3,1:2,2:1}.get(i,1))
            cycle_dmg = ev * copies
            cycle_total += cycle_dmg; cycle_cards += copies
            row = [ident["name"], sk["name"], TYPE_CN[sk["dmg_type"]], SIN_CN[sk["sin"]],
                   base, "+".join(map(str, coins)), len(coins), copies, round(ev,1), round(sd,1),
                   f"{lo:.0f}~{hi:.0f}", round(cycle_dmg,1), tag]
            self._report.append(row)
            self.tree.insert("", "end", values=row[1:], tags=("odd",) if i % 2 else ())
        # DoT 结算
        dot_total, dot_detail, hits = dot_cycle_damage(ident["skills"], enemy_flips=e["enemy_flips"])
        # E.G.O 结算（资源限制次数，替代最弱技能的机会成本）
        ego_txt = ""
        egos = ident.get("egos", [])
        if egos:
            income = sin_income(ident["skills"])
            base_skills = [(r[8], r[1]) for r in self._report]  # (ev, name)
            cheapest_ev = min(base_skills)[0] if base_skills else 0
            for ego in egos:
                mult_e = damage_multiplier(e["phys"], e["sin_res"], ident.get("offense_level", 50),
                                           e["def_lv"], e["power_up"], e["dmg_up"], e["fragile"])
                ev_e, _, _, _ = summarize(skill_damage_dist(ego["base_power"], ego["coin_power"], p, mult_e, e["crit"], CFG["crit_mult"]))
                uses = ego_uses(ego, income)
                net = uses * (ev_e - cheapest_ev)
                cycle_total += net
                ego_txt += f" ｜ {ego['name']}×{uses}次: 期望{ev_e:.1f}，净贡献{net:+.1f}"
        if dot_total:
            cycle_total += dot_total
        dpt = cycle_total / cycle_cards if cycle_cards else 0
        dot_txt = f" ｜ DoT循环 {dot_total:.0f}" if dot_total else ""
        self.status.set(f"理智{e['sanity']}→正面率{p:.0%} ｜ {cycle_cards}张一循环总伤 {cycle_total:.1f}{dot_txt}{ego_txt} ｜ DPT {dpt:.1f}")
        self._cycle = (ident["name"], cycle_cards, round(cycle_total,1), round(dpt,1))

    def clash(self):
        sel = self.tree.selection()
        if len(sel) != 2:
            messagebox.showinfo("提示", "先在表格里按住 Ctrl 选中两个技能再拼点"); return
        e = self.env(); p = heads_prob(e["sanity"]); ident = self.current_ident()
        pair = []
        for s in sel:
            name = self.tree.item(s)["values"][0]
            sk = next(x for x in ident["skills"] if x["name"] == name)
            pair.append((name, *resolve_cond(sk, e["charge"])[:2]))
        (na, ba, ca), (nb, bb, cb) = pair
        wr, rem = clash_sim((ba, ca), (bb, cb), p, p)
        messagebox.showinfo("拼点模拟(1万次)",
            f"{na}  vs  {nb}\n\n{na} 胜率：{wr:.1f}%\n胜时剩余硬币期望：{rem:.2f}")

    def export(self):
        if not getattr(self, "_report", None):
            messagebox.showinfo("提示", "先点「计算伤害期望」"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="limbus_dps报告.xlsx")
        if not path: return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "DPS期望表"
        heads = ["人格","技能","类型","罪孽","基础威力","硬币威力","硬币数","卡组份数","期望伤害","标准差","范围","循环贡献","条件"]
        ws.append(heads)
        for c in range(1, len(heads)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="4472C4")
        for r in self._report: ws.append(r)
        if getattr(self, "_cycle", None):
            ws.append([])
            ws.append([self._cycle[0], f"{self._cycle[1]}张技能一循环总伤", self._cycle[2], "每回合期望(DPT)", self._cycle[3]])
        wb.save(path); self.status.set(f"已导出：{path}")

    # ---- 自定义技能组（含强化技能） ----
    def custom_deck_dialog(self):
        """自定义技能组：统一表单 + 技能库 + 卡组表格。
        交互逻辑：
          · 左侧技能库选人，点「← 载入到表单」把数值填入表单
          · 下方卡组表格选中某行，同样载入表单
          · 表单改完后：「添加为新行」追加 / 「更新选中行」覆盖 / 「删除选中行」
          · E.G.O 勾选后出现消耗与属性字段
        """
        win = tk.Toplevel(self); win.title("自定义技能组 · 循环伤害计算"); win.geometry("980x640")
        deck = []  # [{name, base_power, coin_power, dmg_type, sin, copies, dot?, is_ego?, cost?}]

        # ===== 左：技能库 =====
        left = ttk.LabelFrame(win, text="① 技能库（从已有人格中选）"); left.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        ident_v = tk.StringVar()
        icb = ttk.Combobox(left, textvariable=ident_v, state="readonly", width=36,
                           values=[i["name"] for i in self.idents])
        icb.pack(fill="x", padx=4, pady=4); icb.current(0)
        slb = tk.Listbox(left, width=48, height=20); slb.pack(fill="both", expand=True, padx=4, pady=4)

        def cur_ident():
            return self.idents[icb.current()]

        def fill_lib(*_):
            slb.delete(0, "end")
            for sk in cur_ident()["skills"]:
                slb.insert("end", "{} ｜ 基础{} 硬币{} ｜ {}/{}".format(
                    sk["name"], sk["base_power"], "+".join(map(str, sk["coin_power"])),
                    TYPE_CN[sk["dmg_type"]], SIN_CN[sk["sin"]]))
            if cur_ident().get("egos"):
                for ego in cur_ident()["egos"]:
                    slb.insert("end", "☆{} ｜ 基础{} 硬币{} ｜ 消耗{}".format(
                        ego["name"], ego["base_power"], "+".join(map(str, ego["coin_power"])),
                        ",".join(f"{k}:{v}" for k, v in ego.get("cost", {}).items())))
        icb.bind("<<ComboboxSelected>>", fill_lib); fill_lib()

        # ===== 右：卡组表格 + 表单 =====
        right = ttk.Frame(win); right.pack(side="left", fill="both", expand=True, padx=6, pady=6)

        ttk.Label(right, text="② 当前技能组（选中行可载入表单修改）").pack(anchor="w")
        dcols = ("name", "base", "coins", "copies", "dot")
        dtree = ttk.Treeview(right, columns=dcols, show="headings", height=10)
        for c, h, w in zip(dcols, ["名称", "基础威力", "硬币威力", "份数", "DoT/E.G.O"],
                           [180, 70, 120, 50, 150]):
            dtree.heading(c, text=h); dtree.column(c, width=w, anchor="center")
        dtree.pack(fill="both", expand=True, pady=2)

        # ---- 表单 ----
        form = ttk.LabelFrame(right, text="③ 技能参数（载入后修改，或手动填写）")
        form.pack(fill="x", pady=4)
        f_name = tk.StringVar(); f_base = tk.IntVar(value=4); f_coins = tk.StringVar(value="3,3")
        f_copies = tk.IntVar(value=1); f_sin = tk.StringVar(value="gloom"); f_type = tk.StringVar(value="slash")
        f_dot = tk.StringVar(value="无"); f_dotp = tk.IntVar(value=0); f_dotc = tk.IntVar(value=0)
        f_ego = tk.BooleanVar(value=False); f_cost = tk.StringVar(value="wrath:1,sloth:3")
        r = 0
        ttk.Label(form, text="名称").grid(row=r, column=0, sticky="e")
        ttk.Entry(form, textvariable=f_name, width=24).grid(row=r, column=1, columnspan=3, sticky="w")
        ttk.Label(form, text="类型").grid(row=r, column=4, sticky="e")
        ttk.Combobox(form, textvariable=f_type, state="readonly", width=7,
                     values=list(TYPE_CN)).grid(row=r, column=5, sticky="w")
        ttk.Label(form, text="罪孽").grid(row=r, column=6, sticky="e")
        ttk.Combobox(form, textvariable=f_sin, state="readonly", width=8,
                     values=SINS_LIST).grid(row=r, column=7, sticky="w")
        r += 1
        ttk.Label(form, text="基础威力").grid(row=r, column=0, sticky="e")
        ttk.Spinbox(form, from_=0, to=99, textvariable=f_base, width=5).grid(row=r, column=1, sticky="w")
        ttk.Label(form, text="硬币威力(逗号分隔)").grid(row=r, column=2, sticky="e")
        ttk.Entry(form, textvariable=f_coins, width=12).grid(row=r, column=3, sticky="w")
        ttk.Label(form, text="份数").grid(row=r, column=4, sticky="e")
        ttk.Spinbox(form, from_=1, to=99, textvariable=f_copies, width=4).grid(row=r, column=5, sticky="w")
        ttk.Checkbutton(form, text="E.G.O", variable=f_ego).grid(row=r, column=6, sticky="w")
        ttk.Entry(form, textvariable=f_cost, width=16).grid(row=r, column=7, sticky="w")
        r += 1
        ttk.Label(form, text="DoT").grid(row=r, column=0, sticky="e")
        ttk.Combobox(form, textvariable=f_dot, state="readonly", width=6,
                     values=["无", "流血", "燃烧", "沉潜"]).grid(row=r, column=1, sticky="w")
        ttk.Label(form, text="强度").grid(row=r, column=2, sticky="e")
        ttk.Spinbox(form, from_=0, to=99, textvariable=f_dotp, width=4).grid(row=r, column=3, sticky="w")
        ttk.Label(form, text="次数").grid(row=r, column=4, sticky="e")
        ttk.Spinbox(form, from_=0, to=99, textvariable=f_dotc, width=4).grid(row=r, column=5, sticky="w")
        ttk.Label(form, text="(E.G.O行填消耗 罪孽:数量)").grid(row=r, column=6, columnspan=2, sticky="w")

        def parse_coins(s):
            coins = [float(x) for x in s.replace("，", ",").split(",") if x.strip()]
            return [int(x) if x == int(x) else x for x in coins]

        def parse_cost(s):
            cost = {}
            for part in s.replace("，", ",").split(","):
                if ":" in part:
                    k, v = part.split(":")
                    cost[k.strip()] = int(v)
            return cost

        def read_form():
            """从表单读取并校验，返回技能dict；失败抛异常"""
            coins = parse_coins(f_coins.get())
            if not coins:
                raise ValueError("硬币威力不能为空，格式如 4,4,4")
            if not f_name.get().strip():
                raise ValueError("名称不能为空")
            entry = {"name": f_name.get().strip(), "base_power": f_base.get(),
                     "coin_power": coins, "dmg_type": f_type.get(), "sin": f_sin.get(),
                     "copies": f_copies.get()}
            if f_dot.get() != "无" and f_dotp.get() > 0 and f_dotc.get() > 0:
                entry["dot"] = {"type": DOT_CN_REV[f_dot.get()],
                                "potency": f_dotp.get(), "count": f_dotc.get()}
            if f_ego.get():
                entry["is_ego"] = True
                entry["cost"] = parse_cost(f_cost.get())
            return entry

        def load_to_form(s):
            f_name.set(s["name"]); f_base.set(s["base_power"])
            f_coins.set(",".join(map(str, s["coin_power"])))
            f_copies.set(s.get("copies", 1)); f_sin.set(s["sin"]); f_type.set(s["dmg_type"])
            if s.get("dot"):
                d = s["dot"]
                f_dot.set(DOT_CN.get(d["type"], "无")); f_dotp.set(d["potency"]); f_dotc.set(d["count"])
            else:
                f_dot.set("无"); f_dotp.set(0); f_dotc.set(0)
            f_ego.set(bool(s.get("is_ego")))
            f_cost.set(",".join(f"{k}:{v}" for k, v in s.get("cost", {}).items()) or "wrath:1,sloth:3")

        def refresh_deck():
            dtree.delete(*dtree.get_children())
            for s in deck:
                extra = ""
                if s.get("is_ego"):
                    extra = "E.G.O 消耗[" + ",".join(f"{k}:{v}" for k, v in s.get("cost", {}).items()) + "]"
                elif s.get("dot"):
                    d = s["dot"]
                    extra = f"{DOT_CN.get(d['type'], d['type'])}{d['potency']}强{d['count']}次"
                dtree.insert("", "end", values=(s["name"], s["base_power"],
                             "+".join(map(str, s["coin_power"])), s.get("copies", 1), extra))

        # ---- 交互 ----
        def lib_to_form():
            sel = slb.curselection()
            if not sel:
                messagebox.showinfo("提示", "先在左侧技能库选中一行", parent=win); return
            ident = cur_ident(); idx = sel[0]
            n_skills = len(ident["skills"])
            s = ident["skills"][idx] if idx < n_skills else ident["egos"][idx - n_skills]
            load_to_form(s)

        def deck_to_form(event=None):
            sel = dtree.selection()
            if sel:
                load_to_form(deck[dtree.index(sel[0])])

        def add_row():
            try:
                deck.append(read_form()); refresh_deck()
            except Exception as ex:
                messagebox.showerror("无法添加", str(ex), parent=win)

        def update_row():
            sel = dtree.selection()
            if not sel:
                messagebox.showinfo("提示", "先在卡组表格选中要更新的行", parent=win); return
            try:
                deck[dtree.index(sel[0])] = read_form()
                refresh_deck()
            except Exception as ex:
                messagebox.showerror("无法更新", str(ex), parent=win)

        def del_row():
            for i in reversed([dtree.index(s) for s in dtree.selection()]):
                deck.pop(i)
            refresh_deck()

        dtree.bind("<<TreeviewSelect>>", deck_to_form)

        bf = ttk.Frame(right); bf.pack(fill="x", pady=2)
        ttk.Button(bf, text="← 载入库中选中", command=lib_to_form).pack(side="left", padx=2)
        ttk.Button(bf, text="添加为新行", command=add_row).pack(side="left", padx=2)
        ttk.Button(bf, text="更新选中行", command=update_row).pack(side="left", padx=2)
        ttk.Button(bf, text="删除选中行", command=del_row).pack(side="left", padx=2)

        # ---- 结果区 ----
        ttk.Label(right, text="④ 计算结果").pack(anchor="w", pady=(6, 0))
        rlb = tk.Listbox(right, width=52, height=7); rlb.pack(fill="both", expand=True, pady=2)
        result_v = tk.StringVar(value="")
        ttk.Label(right, textvariable=result_v, font=("微软雅黑", 11, "bold"),
                  foreground="#1F4E79").pack(anchor="w", pady=2)

        def calc_deck():
            if not deck:
                return
            e = self.env(); p = heads_prob(e["sanity"])
            ident = self.current_ident()
            mult = damage_multiplier(e["phys"], e["sin_res"], ident.get("offense_level", 50),
                                     e["def_lv"], e["power_up"], e["dmg_up"], e["fragile"])
            total = 0.0; cards = 0; lines = []; evs = []
            for s in deck:
                if s.get("is_ego"):
                    continue
                ev, sd, lo, hi = summarize(skill_damage_dist(
                    s["base_power"], s["coin_power"], p, mult, e["crit"], CFG["crit_mult"]))
                total += ev * s["copies"]; cards += s["copies"]; evs.append(ev)
                lines.append(f"{s['name']} ×{s['copies']}：期望 {ev:.1f}，贡献 {ev * s['copies']:.1f}")
            dot_total, dot_detail, hits = dot_cycle_damage(deck, enemy_flips=e["enemy_flips"])
            if dot_total:
                total += dot_total
                lines.append(f"—— DoT循环合计：{dot_total:.0f}（{'、'.join(dot_detail)}）")
            egos = [s for s in deck if s.get("is_ego")]
            if egos:
                income = sin_income(deck)
                cheapest = min(evs) if evs else 0
                for ego in egos:
                    ev_e = summarize(skill_damage_dist(
                        ego["base_power"], ego["coin_power"], p, mult, e["crit"], CFG["crit_mult"]))[0]
                    uses = ego_uses(ego, income)
                    net = uses * (ev_e - cheapest)
                    total += net
                    lines.append(f"☆{ego['name']}：期望 {ev_e:.1f}，资源可发动 {uses} 次/循环，"
                                 f"净贡献 {net:+.1f}（替代最弱牌 {cheapest:.1f}）")
            dpt = total / cards if cards else 0
            result_v.set(f"{cards}张一循环 总伤期望 {total:.1f} ｜ 每回合(DPT) {dpt:.1f}")
            rlb.delete(0, "end")
            for line in lines:
                rlb.insert("end", line)

        def save_as_identity():
            if not deck:
                return
            name = simpledialog.askstring("保存", "组合名称:", parent=win)
            if not name:
                return
            skills = [dict(s) for s in deck if not s.get("is_ego")]
            egos = [dict(s) for s in deck if s.get("is_ego")]
            ident = {"name": name, "note": "自定义技能组", "offense_level": 50, "skills": skills}
            if egos:
                ident["egos"] = egos
            self.idents.append(ident)
            with open(IDENT_FILE, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            self.reload()
            result_v.set(f"已保存为人格：{name}")

        bf2 = ttk.Frame(right); bf2.pack(fill="x", pady=2)
        ttk.Button(bf2, text="计算循环伤害", command=calc_deck).pack(side="left", padx=2)
        ttk.Button(bf2, text="保存为人格", command=save_as_identity).pack(side="left", padx=2)


    # ---- Wiki 数据获取 ----
    def wiki_dialog(self):
        import wiki_fetch
        win = tk.Toplevel(self); win.title("从 Wiki 获取人格数据"); win.geometry("560x480")
        ttk.Label(win, text="搜索(名称/罪人):").pack(anchor="w", padx=8, pady=(8,0))
        mode_v = tk.StringVar(value="ident")
        mf = ttk.Frame(win); mf.pack(fill="x", padx=8)
        ttk.Radiobutton(mf, text="人格", variable=mode_v, value="ident", command=lambda: load()).pack(side="left")
        ttk.Radiobutton(mf, text="E.G.O", variable=mode_v, value="ego", command=lambda: load()).pack(side="left")
        search_v = tk.StringVar()
        ent = ttk.Entry(win, textvariable=search_v, width=40); ent.pack(fill="x", padx=8, pady=4)
        lb = tk.Listbox(win, width=70, height=18); lb.pack(fill="both", expand=True, padx=8, pady=4)
        stat = tk.StringVar(value="正在加载人格列表…")
        ttk.Label(win, textvariable=stat).pack(anchor="w", padx=8)
        self._wiki_rows = []
        self._all_wiki_rows = []

        def fill(rows):
            lb.delete(0, "end"); self._wiki_rows = rows
            for name, _ in rows: lb.insert("end", name)
            stat.set(f"共 {len(rows)} 个结果，双击或点「导入」")

        def load(refresh=False):
            stat.set("正在从 Wiki 数据源下载…")
            mode = mode_v.get()
            def work():
                try:
                    if mode == "ident":
                        rows = wiki_fetch.list_identities(refresh=refresh)
                    else:
                        rows = wiki_fetch.list_egos(refresh=refresh)
                    self._all_wiki_rows = rows
                    self.after(0, lambda: (fill(rows), ent.focus()))
                except Exception as ex:
                    self.after(0, lambda: stat.set(f"下载失败：{ex}"))
            import threading; threading.Thread(target=work, daemon=True).start()

        def on_search(*_):
            q = search_v.get().lower()
            fill([r for r in getattr(self, "_all_wiki_rows", []) if q in r[0].lower()])

        def do_import(event=None):
            sel = lb.curselection()
            if not sel: return
            name, item_id = self._wiki_rows[sel[0]]
            try:
                if mode_v.get() == "ego":
                    ego = wiki_fetch.convert_ego(item_id)
                    ident = self.current_ident()
                    ident.setdefault("egos", [])
                    if any(g["name"] == ego["name"] for g in ident["egos"]):
                        messagebox.showinfo("提示", "该E.G.O已挂在当前人格下", parent=win); return
                    ident["egos"].append(ego)
                    with open(IDENT_FILE, "w", encoding="utf-8") as f:
                        json.dump(self.data, f, ensure_ascii=False, indent=2)
                    stat.set(f"已导入 {ego['name']} → 挂到「{ident['name']}」（消耗 {ego['cost']}）")
                    return
                ident = wiki_fetch.convert_identity(item_id)
                if any(i["name"] == ident["name"] for i in self.idents):
                    messagebox.showinfo("提示", "该人格已存在", parent=win); return
                self.idents.append(ident)
                with open(IDENT_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.data, f, ensure_ascii=False, indent=2)
                self.reload()
                self.ident_cb.current(len(self.idents)-1); self.refresh_skills()
                stat.set(f"已导入：{ident['name']}（{len(ident['skills'])}个技能）")
            except Exception as ex:
                messagebox.showerror("导入失败", str(ex), parent=win)

        ent.bind("<KeyRelease>", on_search)
        lb.bind("<Double-Button-1>", do_import)
        bf = ttk.Frame(win); bf.pack(fill="x", padx=8, pady=6)
        ttk.Button(bf, text="导入选中人格", command=do_import).pack(side="left", padx=4)
        ttk.Button(bf, text="强制刷新缓存", command=lambda: load(True)).pack(side="left", padx=4)
        load()

    # ---- 人格数据管理 ----
    def open_json(self):
        os.startfile(IDENT_FILE) if os.name == "nt" else subprocess.call(["open", IDENT_FILE])

    def reload(self):
        try:
            self.load_data()
            self.ident_cb["values"] = [i["name"] for i in self.idents]
            self.ident_cb.current(0); self.refresh_skills()
            self.status.set("数据已重新加载")
        except Exception as ex:
            messagebox.showerror("加载失败", str(ex))

    def add_identity_dialog(self):
        win = tk.Toplevel(self); win.title("新增人格"); win.geometry("520x420")
        ttk.Label(win, text="人格名称:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        name_v = tk.StringVar(); ttk.Entry(win, textvariable=name_v, width=30).grid(row=0, column=1)
        ttk.Label(win, text="攻击等级:").grid(row=1, column=0, sticky="e", padx=4)
        lv_v = tk.IntVar(value=50); ttk.Spinbox(win, from_=1, to=90, textvariable=lv_v, width=6).grid(row=1, column=1, sticky="w")
        ttk.Label(win, text='技能(JSON数组)，例:\n[{"name":"S1","base_power":4,"coin_power":[3,3],\n"dmg_type":"slash","sin":"lust"}]',
                  justify="left").grid(row=2, column=0, columnspan=2, padx=6, sticky="w")
        txt = tk.Text(win, width=60, height=14); txt.grid(row=3, column=0, columnspan=2, padx=6, pady=4)
        txt.insert("1.0", '[\n  {"name":"S1","base_power":4,"coin_power":[3,3],"dmg_type":"slash","sin":"lust"},\n'
                          '  {"name":"S2","base_power":5,"coin_power":[4,4,4],"dmg_type":"pierce","sin":"wrath"},\n'
                          '  {"name":"S3","base_power":6,"coin_power":[5,5,5,5],"dmg_type":"blunt","sin":"gloom"}\n]')
        def save():
            try:
                skills = json.loads(txt.get("1.0", "end"))
                assert name_v.get().strip(), "人格名称不能为空"
                for sk in skills:
                    for k in ("name","base_power","coin_power","dmg_type","sin"): assert k in sk, f"技能缺少字段 {k}"
                    assert sk["dmg_type"] in TYPE_CN and sk["sin"] in SIN_CN, "dmg_type/sin 取值非法"
                self.idents.append({"name": name_v.get().strip(), "note": "自定义",
                                    "offense_level": lv_v.get(), "skills": skills})
                with open(IDENT_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.data, f, ensure_ascii=False, indent=2)
                self.reload(); win.destroy()
            except Exception as ex:
                messagebox.showerror("保存失败", str(ex), parent=win)
        ttk.Button(win, text="保存", command=save).grid(row=4, column=1, sticky="e", padx=6, pady=6)


if __name__ == "__main__":
    App().mainloop()
