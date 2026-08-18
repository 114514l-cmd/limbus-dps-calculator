# -*- coding: utf-8 -*-
"""
边狱巴士（Limbus Company）人格 DPS 计算器
================================================
核心模型（均为社区共识近似，参数可调）：
1. 硬币机制：技能伤害 = Σ 每个硬币的命中伤害
   每硬币命中伤害 = (基础威力 + 硬币威力[正面时]) × 各类乘区
   正面概率 p = 0.5 + 理智度×1%（限制在 5%~95%）
2. 拼点（Clash）：双方掷全部剩余硬币比威力，败方损失1硬币，循环至一方无硬币；
   胜者用剩余硬币造成伤害（蒙特卡洛模拟）
3. 乘区：
   - 物理抗性(斩/突/打)：致命1.5 / 普通1.0 / 耐受0.5 / 无效0
   - 罪孽抗性(七宗罪)：同上
   - 攻防等级差：每级 ±3%（可调）
   - 暴击：基础暴击率5%（呼吸/充能等可加），暴击伤害×1.2（可调）
   - 增益：攻击威力提升%、伤害提升%、敌方脆弱%
4. 输出：每个技能的期望伤害/标准差/最小/最大、DPS 排行、拼点胜率模拟

用法：
    python dps_calc.py                      # 全人格技能期望伤害表
    python dps_calc.py --sanity 30          # 指定理智度(影响正面率)
    python dps_calc.py --clash              # 进入拼点模拟模式
    python dps_calc.py --excel              # 导出 Excel 报告
"""
import json, os, sys, math, random, argparse, itertools

BASE = os.path.dirname(os.path.abspath(__file__))

# ---------------- 全局参数（可按版本调整） ----------------
CFG = {
    "sanity_to_heads": 0.01,     # 每点理智对正面概率的影响
    "heads_min": 0.05, "heads_max": 0.95,
    "crit_chance": 0.05,         # 基础暴击率
    "crit_mult": 1.2,            # 暴击倍率（社区口径1.2x，部分版本/被动会提高）
    "level_diff_coef": 0.03,     # 攻防等级每差1级的伤害变化
    "resist_table": {            # 抗性映射
        "fatal": 1.5, "normal": 1.0, "endured": 0.5, "ineffective": 0.0,
    },
}

PHYS_TYPES = ["slash", "pierce", "blunt"]
SINS = ["wrath", "lust", "sloth", "gluttony", "gloom", "pride", "envy"]
TYPE_CN = {"slash": "斩击", "pierce": "突刺", "blunt": "打击"}
SIN_CN = {"wrath":"愤怒","lust":"色欲","sloth":"怠惰","gluttony":"暴食",
          "gloom":"忧郁","pride":"傲慢","envy":"嫉妒"}


def heads_prob(sanity):
    return min(max(0.5 + sanity * CFG["sanity_to_heads"], CFG["heads_min"]), CFG["heads_max"])


def resolve_cond(skill, charge):
    """应用条件增益，返回 (base_power, coin_power_list, 触发说明)"""
    base = skill["base_power"]; coins = list(skill["coin_power"]); tag = ""
    c = skill.get("cond")
    if c and charge >= c.get("charge_at_least", 10**9):
        base += c.get("base_power_add", 0)
        coins = [x + c.get("coin_power_add", 0) for x in coins]
        tag = f"(充能≥{c['charge_at_least']}已触发)"
    return base, coins, tag


def damage_multiplier(phys_resist, sin_resist, off_lv, def_lv, power_up, dmg_up, fragile):
    m = CFG["resist_table"][phys_resist] * CFG["resist_table"][sin_resist]
    m *= 1 + CFG["level_diff_coef"] * (off_lv - def_lv)
    m *= (1 + power_up / 100) * (1 + dmg_up / 100) * (1 + fragile / 100)
    return m


def skill_damage_dist(base, coins, p, mult, crit_chance, crit_mult):
    """枚举全部硬币结果，返回伤害分布 [(damage, prob)]"""
    dist = {}
    n = len(coins)
    for outcomes in itertools.product([0, 1], repeat=n):
        prob = 1.0; total = 0.0
        for cp, heads in zip(coins, outcomes):
            prob *= p if heads else (1 - p)
            hit = base + (cp if heads else 0)
            # 每个硬币独立判定暴击的期望处理：直接乘期望系数
            hit *= 1 + crit_chance * (crit_mult - 1)
            total += hit
        total *= mult
        dist[round(total, 2)] = dist.get(round(total, 2), 0) + prob
    return sorted(dist.items())


def summarize(dist):
    ev = sum(d * q for d, q in dist)
    var = sum((d - ev) ** 2 * q for d, q in dist)
    lo = dist[0][0]; hi = dist[-1][0]
    return ev, math.sqrt(var), lo, hi


def clash_sim(a, b, p_a, p_b, rounds=20000):
    """蒙特卡洛拼点：a,b 为 (base, coins) 元组。返回 (a胜率%, a剩余硬币期望)"""
    a_base, a_coins = a; b_base, b_coins = b
    win = 0; remain_sum = 0
    for _ in range(rounds):
        ca = list(a_coins); cb = list(b_coins)
        while ca and cb:
            pa = a_base + sum(cp for cp in ca if random.random() < p_a)
            pb = b_base + sum(cp for cp in cb if random.random() < p_b)
            if pa > pb: cb.pop()
            elif pb > pa: ca.pop()
            else: continue  # 平起重掷
        if ca:
            win += 1; remain_sum += len(ca)
    return win / rounds * 100, remain_sum / rounds


# ---------------- E.G.O / DoT ----------------
def sin_income(deck):
    inc = {}
    for s in deck:
        if s.get("is_ego"): continue
        inc[s["sin"]] = inc.get(s["sin"], 0) + s.get("copies", 1)
    return inc

def ego_uses(ego, income, turns=6):
    cost = ego.get("cost", {})
    if not cost: return 1
    return min(min(income.get(s, 0) // n for s, n in cost.items()), turns)

DOT_CN = {"bleed": "流血", "burn": "燃烧", "sinking": "沉潜"}
def dot_cycle_damage(deck, turns=6, enemy_flips=12):
    """流血=中招者拼点/用硬币时触发；沉潜=被命中触发；燃烧=回合结束触发"""
    hits = sum(s.get("copies", 1) * len(s["coin_power"]) for s in deck if not s.get("is_ego"))
    total = 0.0
    for s in deck:
        dot = s.get("dot")
        if not dot or s.get("is_ego"): continue
        p_, c_ = dot.get("potency", 0), dot.get("count", 0)
        for _ in range(s.get("copies", 1)):
            if dot["type"] == "burn":
                total += p_ * min(c_, turns)
            elif dot["type"] == "bleed":
                total += p_ * min(c_, enemy_flips)
            else:
                total += p_ * min(c_, hits)
    return total


# ---------------- 主流程 ----------------
def main():
    ap = argparse.ArgumentParser(description="边狱巴士人格DPS计算器")
    ap.add_argument("--sanity", type=int, default=0, help="理智度(-45~45)，影响硬币正面率")
    ap.add_argument("--charge", type=int, default=0, help="充能层数（触发条件增益）")
    ap.add_argument("--def-lv", type=int, default=50, help="敌方防御等级")
    ap.add_argument("--phys", default="normal", choices=list(CFG["resist_table"]), help="敌方物理抗性")
    ap.add_argument("--sin-res", default="normal", choices=list(CFG["resist_table"]), help="敌方罪孽抗性")
    ap.add_argument("--power-up", type=float, default=0, help="攻击威力提升%%")
    ap.add_argument("--dmg-up", type=float, default=0, help="伤害提升%%")
    ap.add_argument("--fragile", type=float, default=0, help="敌方脆弱%%")
    ap.add_argument("--enemy-flips", type=int, default=12, help="敌方每循环拼点/硬币行动数(影响流血触发)")
    ap.add_argument("--crit-chance", type=float, default=None, help="暴击率(0-1)，默认读配置")
    ap.add_argument("--clash", action="store_true", help="拼点模拟模式（各人格S3互拼）")
    ap.add_argument("--excel", action="store_true", help="导出Excel报告")
    args = ap.parse_args()

    data = json.load(open(os.path.join(BASE, "identities.json"), encoding="utf-8"))
    p = heads_prob(args.sanity)
    cc = args.crit_chance if args.crit_chance is not None else CFG["crit_chance"]

    print(f"\n===== 环境设定 =====")
    print(f"理智度 {args.sanity} → 硬币正面率 {p:.0%} ｜ 充能 {args.charge} ｜ 敌方防御等级 {args.def_lv}")
    print(f"敌方物理抗性 {args.phys}({CFG['resist_table'][args.phys]}x) ｜ 罪孽抗性 {args.sin_res}({CFG['resist_table'][args.sin_res]}x)")
    print(f"暴击率 {cc:.0%}，暴击倍率 {CFG['crit_mult']}x\n")

    report = []
    for ident in data["identities"]:
        print(f"■ {ident['name']}（{ident.get('note','')}，攻击等级{ident.get('offense_level',50)}）")
        print(f"  {'技能':<14}{'基础':>4}{'硬币':>10}{'份数':>4}{'期望伤害':>9}{'标准差':>7}{'范围':>12}  备注")
        cycle_total = 0.0; cycle_cards = 0
        for i, sk in enumerate(ident["skills"]):
            base, coins, tag = resolve_cond(sk, args.charge)
            mult = damage_multiplier(args.phys, args.sin_res,
                                     ident.get("offense_level", 50), args.def_lv,
                                     args.power_up, args.dmg_up, args.fragile)
            dist = skill_damage_dist(base, coins, p, mult, cc, CFG["crit_mult"])
            ev, sd, lo, hi = summarize(dist)
            copies = sk.get("copies", {0:3,1:2,2:1}.get(i,1))
            cycle_total += ev * copies; cycle_cards += copies
            coin_str = "+".join(str(c) for c in coins)
            print(f"  {sk['name']:<14}{base:>4}{coin_str:>10}{copies:>4}{ev:>9.1f}{sd:>7.1f}{f'{lo:.0f}~{hi:.0f}':>12}  {tag}")
            report.append([ident["name"], sk["name"], TYPE_CN[sk["dmg_type"]], SIN_CN[sk["sin"]],
                           base, coin_str, len(coins), round(ev,1), round(sd,1), round(lo,1), round(hi,1), tag])
        dpt = cycle_total / cycle_cards if cycle_cards else 0
        # DoT
        dot_total = dot_cycle_damage(ident["skills"], enemy_flips=args.enemy_flips)
        # E.G.O
        ego_txt = ""
        for ego in ident.get("egos", []):
            income = sin_income(ident["skills"])
            mult_e = damage_multiplier(args.phys, args.sin_res,
                                       ident.get("offense_level", 50), args.def_lv,
                                       args.power_up, args.dmg_up, args.fragile)
            dist_e = skill_damage_dist(ego["base_power"], ego["coin_power"], p, mult_e, cc, CFG["crit_mult"])
            ev_e = summarize(dist_e)[0]
            cheapest = min(r[7] for r in report if r[0] == ident["name"])
            uses = ego_uses(ego, income)
            net = uses * (ev_e - cheapest)
            cycle_total += net
            ego_txt += f" ｜ {ego['name']}×{uses}: 期望{ev_e:.1f} 净{net:+.1f}"
        if dot_total:
            cycle_total += dot_total
        dpt = cycle_total / cycle_cards if cycle_cards else 0
        dot_txt = f"，DoT循环 {dot_total:.0f}" if dot_total else ""
        print(f"  → {cycle_cards}张技能一循环：总伤期望 {cycle_total:.1f}{dot_txt}{ego_txt}，DPT {dpt:.1f}")
        print()

    # DPS 排行
    print("===== 期望伤害排行（全部技能） =====")
    for i, r in enumerate(sorted(report, key=lambda x: -x[7])[:8], 1):
        print(f"  {i}. {r[0]} · {r[1]}：期望 {r[7]}（{r[6]}硬币，{r[3]}）")

    # 拼点模拟
    if args.clash:
        print("\n===== 拼点模拟（各人格S3互拼，2万次蒙特卡洛） =====")
        s3 = []
        for ident in data["identities"]:
            sk = ident["skills"][-1]
            base, coins, _ = resolve_cond(sk, args.charge)
            s3.append((ident["name"], base, coins))
        for i in range(len(s3)):
            for j in range(i + 1, len(s3)):
                na, ba, ca = s3[i]; nb, bb, cb = s3[j]
                wr, rem = clash_sim((ba, ca), (bb, cb), p, p)
                print(f"  {na} vs {nb}：胜率 {wr:.1f}% ｜ 胜时剩余硬币期望 {rem:.2f}")

    # Excel 导出
    if args.excel:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "DPS期望表"
        heads = ["人格","技能","伤害类型","罪孽","基础威力","硬币威力","硬币数",
                 "期望伤害","标准差","最小","最大","条件触发"]
        ws.append(heads)
        for c in range(1, len(heads)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="4472C4")
        for r in report: ws.append(r)
        ws2 = wb.create_sheet("环境参数")
        for k, v in [("理智度", args.sanity), ("正面率", f"{p:.0%}"), ("充能", args.charge),
                     ("敌方防御等级", args.def_lv), ("物理抗性", args.phys), ("罪孽抗性", args.sin_res),
                     ("暴击率", cc), ("暴击倍率", CFG["crit_mult"]),
                     ("等级差系数", CFG["level_diff_coef"])]:
            ws2.append([k, v])
        out = os.path.join(BASE, "limbus_dps报告.xlsx")
        wb.save(out); print(f"\nExcel 已导出：{out}")


if __name__ == "__main__":
    main()
